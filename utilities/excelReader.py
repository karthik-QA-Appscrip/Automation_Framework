from openpyxl import load_workbook
from utilities.configReader import ConfigReader


class ExcelReader:

    def __init__(self, file_path, sheet_name):
        config = ConfigReader()
        self.workbook = load_workbook(config.get_excel_path())
        self.sheet = self.workbook[config.get_sheet_name()]

    def get_data(self):
        data = []

        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            data.append(row)

        return data